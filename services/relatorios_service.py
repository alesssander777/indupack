"""Motor analítico dos relatórios gerenciais — SQLite (apontamentos) + estado JSON (paradas)."""
from __future__ import annotations

import io
import logging
import os
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from sqlalchemy import select
from sqlalchemy.orm import Session

from database.database import SessionLocal
from database.models import Apontamento
from storage.state import dados_maquinas
from reportlab.platypus.flowables import Flowable

_ROOT = Path(__file__).resolve().parent.parent

_log = logging.getLogger(__name__)


@dataclass
class FiltrosRelatorio:
    inicio: datetime
    fim: datetime
    maquina_id: int | None = None
    operador: str = ""
    turno: str = ""
    produto: str = ""
    setor: str = ""

    def maquinas_permitidas(self) -> list[int] | None:
        """None = todas as máquinas ativas; senão lista filtrada por id e/ou setor."""
        setor_f = (self.setor or "").strip().lower()
        out: list[int] = []
        for mid, m in sorted(dados_maquinas.items()):
            if m.get("ativo", True) is False:
                continue
            if self.maquina_id is not None and int(mid) != int(self.maquina_id):
                continue
            if setor_f:
                st = str(m.get("setor") or "").strip().lower()
                if setor_f not in st:
                    continue
            out.append(int(mid))
        return out if out else ([] if self.maquina_id or setor_f else None)


def _day_bounds(d: date) -> tuple[datetime, datetime]:
    start = datetime.combine(d, datetime.min.time())
    end = datetime.combine(d, datetime.max.time())
    return start, end


def parse_filtros(
    inicio_s: str | None,
    fim_s: str | None,
    maquina: str | None = None,
    operador: str | None = None,
    turno: str | None = None,
    produto: str | None = None,
    setor: str | None = None,
) -> FiltrosRelatorio:
    today = date.today()
    try:
        di = date.fromisoformat((inicio_s or "").strip() or today.isoformat())
    except ValueError:
        di = today
    try:
        df = date.fromisoformat((fim_s or "").strip() or today.isoformat())
    except ValueError:
        df = today
    if df < di:
        df = di
    t0, _ = _day_bounds(di)
    _, t1 = _day_bounds(df)
    mid: int | None = None
    if maquina not in (None, "", "0"):
        try:
            mid = int(maquina)
        except (TypeError, ValueError):
            mid = None
    return FiltrosRelatorio(
        inicio=t0,
        fim=t1,
        maquina_id=mid,
        operador=(operador or "").strip(),
        turno=(turno or "").strip(),
        produto=(produto or "").strip(),
        setor=(setor or "").strip(),
    )


def _nome_maquina(mid: int) -> str:
    m = dados_maquinas.get(mid) or {}
    n = str(m.get("nome") or "").strip()
    return n or f"Máquina {mid}"


def _setor_maquina(mid: int) -> str:
    m = dados_maquinas.get(mid) or {}
    return str(m.get("setor") or "").strip()


def _meta_maquina(mid: int) -> int:
    m = dados_maquinas.get(mid) or {}
    try:
        return max(1, int(m.get("meta") or 1000))
    except (TypeError, ValueError):
        return 1000


def _baselines(session: Session, t0: datetime, maquinas: list[int]) -> dict[int, int]:
    out: dict[int, int] = {}
    for mid in maquinas:
        row = session.scalar(
            select(Apontamento)
            .where(Apontamento.maquina == mid, Apontamento.horario < t0)
            .order_by(Apontamento.horario.desc())
            .limit(1)
        )
        if row is not None:
            out[mid] = int(row.quantidade or 0)
    return out


def _query_apontamentos(session: Session, flt: FiltrosRelatorio, maquinas: list[int]) -> list[Apontamento]:
    q = select(Apontamento).where(
        Apontamento.horario >= flt.inicio,
        Apontamento.horario <= flt.fim,
        Apontamento.maquina.in_(maquinas),
    )
    op = (flt.operador or "").strip()
    if op:
        q = q.where(Apontamento.operador.ilike(f"%{op}%"))
    tu = (flt.turno or "").strip()
    if tu:
        q = q.where(Apontamento.turno.ilike(f"%{tu}%"))
    pr = (flt.produto or "").strip()
    if pr:
        q = q.where(Apontamento.produto.ilike(f"%{pr}%"))
    q = q.order_by(Apontamento.maquina, Apontamento.horario)
    return list(session.scalars(q).all())


def _delta_events(rows: list[Apontamento], baselines: dict[int, int]) -> list[dict[str, Any]]:
    by_m: dict[int, list[Apontamento]] = defaultdict(list)
    for r in rows:
        by_m[int(r.maquina)].append(r)
    events: list[dict[str, Any]] = []
    for mid, lst in by_m.items():
        prev = baselines.get(mid)
        prev_t: datetime | None = None
        for r in lst:
            q = int(r.quantidade or 0)
            t = r.horario
            if prev is None:
                prev = q
                prev_t = t
                continue
            delta = max(0, q - prev)
            if delta > 0:
                dt_s = 0
                if prev_t is not None:
                    dt_s = max(0, int((t - prev_t).total_seconds()))
                events.append(
                    {
                        "maquina": mid,
                        "maquina_nome": _nome_maquina(mid),
                        "setor": _setor_maquina(mid),
                        "operador": str(r.operador or "").strip(),
                        "turno": str(r.turno or "").strip(),
                        "produto": str(r.produto or "").strip(),
                        "quantidade": delta,
                        "horario": t.isoformat(timespec="seconds"),
                        "status": str(r.status or "").strip(),
                        "tempo_producao_est_s": dt_s,
                    }
                )
            prev = q
            prev_t = t
    events.sort(key=lambda x: x["horario"])
    return events


def _epoch_ms(dt: datetime) -> int:
    return int(dt.timestamp() * 1000)


def coletar_paradas(flt: FiltrosRelatorio, maquinas: list[int]) -> list[dict[str, Any]]:
    t0 = _epoch_ms(flt.inicio)
    t1 = _epoch_ms(flt.fim)
    out: list[dict[str, Any]] = []
    for mid in maquinas:
        m = dados_maquinas.get(mid) or {}
        hist = m.get("historico_paradas")
        if not isinstance(hist, list):
            continue
        for h in hist:
            try:
                ini = int(h.get("inicio_epoch") or 0)
                ret = int(h.get("retorno_epoch") or 0)
            except (TypeError, ValueError):
                continue
            if ret <= 0 or ini <= 0:
                continue
            if ret < t0 or ini > t1:
                continue
            dur = int(h.get("duracao_s") or max(0, (ret - ini) // 1000))
            out.append(
                {
                    "maquina": mid,
                    "maquina_nome": _nome_maquina(mid),
                    "setor": _setor_maquina(mid),
                    "inicio": datetime.fromtimestamp(ini / 1000.0).isoformat(timespec="seconds"),
                    "retorno": datetime.fromtimestamp(ret / 1000.0).isoformat(timespec="seconds"),
                    "duracao_s": dur,
                    "motivo": str(h.get("motivo") or ""),
                    "operador": str(h.get("operador") or "").strip(),
                    "turno": str(h.get("turno") or "").strip(),
                }
            )
    out.sort(key=lambda x: x["inicio"])
    return out


def _fmt_hms(seg: int) -> str:
    seg = max(0, int(seg))
    h, r = divmod(seg, 3600)
    m, s = divmod(r, 60)
    return f"{h:02d}:{m:02d}:{s:02d}"


def _agg_counter(events: list[dict], key: str) -> list[dict[str, Any]]:
    acc: dict[str, int] = defaultdict(int)
    for e in events:
        k = str(e.get(key) or "").strip() or "—"
        acc[k] += int(e.get("quantidade") or 0)
    rows = [{"nome": k, "quantidade": v} for k, v in acc.items()]
    rows.sort(key=lambda x: -x["quantidade"])
    return rows


def montar_relatorio(flt: FiltrosRelatorio) -> dict[str, Any]:
    maquinas = flt.maquinas_permitidas()
    if maquinas is None:
        maquinas = []
        for mid, m in sorted(dados_maquinas.items()):
            if m.get("ativo", True) is False:
                continue
            maquinas.append(int(mid))
    if not maquinas:
        return _empty_payload(flt, "Nenhuma máquina no filtro ou cadastro vazio.")

    with SessionLocal() as session:
        baselines = _baselines(session, flt.inicio, maquinas)
        rows = _query_apontamentos(session, flt, maquinas)
        events = _delta_events(rows, baselines)
    paradas = coletar_paradas(flt, maquinas)

    prod_total = sum(int(e["quantidade"]) for e in events)
    parada_s = sum(int(p["duracao_s"]) for p in paradas)
    span_s = max(1, int((flt.fim - flt.inicio).total_seconds()) + 1)
    n_maq = max(1, len(maquinas))
    wall_s = span_s * n_maq
    disp_proxy = 100.0 * max(0.0, wall_s - parada_s) / wall_s if wall_s else 0.0

    meta_ref = sum(_meta_maquina(mid) for mid in maquinas) * max(1, span_s / 86400.0)
    perf_proxy = min(1.5, prod_total / max(1.0, meta_ref))
    oee_proxy = round((disp_proxy / 100.0) * perf_proxy * 100.0, 1)

    por_maquina = _agg_counter(events, "maquina_nome")
    por_operador = _agg_counter(events, "operador")
    por_turno = _agg_counter(events, "turno")

    paradas_por_maq: dict[int, int] = defaultdict(int)
    for p in paradas:
        paradas_por_maq[int(p["maquina"])] += int(p["duracao_s"])

    comparacao_maquinas = []
    for mid in maquinas:
        nome = _nome_maquina(mid)
        prod_m = sum(int(e["quantidade"]) for e in events if int(e["maquina"]) == mid)
        par_m = paradas_por_maq.get(mid, 0)
        meta = _meta_maquina(mid)
        dias = max(span_s / 86400.0, 0.01)
        ef = round(100.0 * prod_m / max(1, meta * dias), 1)
        comparacao_maquinas.append(
            {
                "maquina_id": mid,
                "nome": nome,
                "setor": _setor_maquina(mid),
                "producao": prod_m,
                "tempo_parado_s": par_m,
                "tempo_parado_fmt": _fmt_hms(par_m),
                "eficiencia_est": ef,
                "meta_referencia": int(meta * dias),
            }
        )
    comparacao_maquinas.sort(key=lambda x: -x["producao"])

    produtos_rank = _agg_counter(events, "produto")[:25]

    detalhado: list[dict[str, Any]] = []
    for e in events:
        tp = int(e.get("tempo_producao_est_s") or 0)
        detalhado.append(
            {
                "produto": e.get("produto") or "—",
                "operador": e.get("operador") or "—",
                "turno": e.get("turno") or "—",
                "maquina": e.get("maquina_nome") or "—",
                "maquina_id": e.get("maquina"),
                "quantidade": e.get("quantidade"),
                "horario": e.get("horario"),
                "tempo_producao": _fmt_hms(tp) if tp else "—",
                "tempo_producao_s": tp,
                "tempo_parado": "—",
                "observacoes": (e.get("status") or "—") or "—",
            }
        )

    paradas_resumo = _agg_counter(
        [{"nome": (p.get("motivo") or "Sem motivo").strip() or "—", "quantidade": int(p["duracao_s"])} for p in paradas],
        "nome",
    )
    for r in paradas_resumo:
        r["motivo"] = r.get("nome") or "—"
        r["duracao_s"] = r.pop("quantidade", 0)
        r["duracao_fmt"] = _fmt_hms(int(r["duracao_s"]))

    perdas_estimadas = 0
    for c in comparacao_maquinas:
        perdas_estimadas += max(0, int(c["meta_referencia"]) - int(c["producao"]))

    return {
        "ok": True,
        "filtros": {
            "inicio": flt.inicio.date().isoformat(),
            "fim": flt.fim.date().isoformat(),
            "maquina_id": flt.maquina_id,
            "operador": flt.operador,
            "turno": flt.turno,
            "produto": flt.produto,
            "setor": flt.setor,
        },
        "notas_metodologia": [
            "Produção calculada a partir dos apontamentos (variação do total acumulado por máquina).",
            "Paradas: histórico persistido no estado da máquina (últimos eventos retomados no período).",
            "Tempo entre apontamentos positivos é estimativa de ritmo; OEE é proxy (disponibilidade × desempenho vs meta).",
        ],
        "kpis": {
            "producao_total": prod_total,
            "total_apontamentos_db": len(rows),
            "eventos_producao": len(events),
            "tempo_parado_total_s": parada_s,
            "tempo_parado_fmt": _fmt_hms(parada_s),
            "disponibilidade_proxy_pct": round(disp_proxy, 1),
            "oee_proxy_pct": oee_proxy,
            "perdas_estimadas_vs_meta": max(0, perdas_estimadas),
            "maquinas_no_escopo": len(maquinas),
        },
        "diario": {
            "por_maquina": por_maquina,
            "por_operador": por_operador,
            "por_turno": por_turno,
            "produtos": produtos_rank[:20],
            "paradas_motivo": paradas_resumo[:15],
            "eficiencia_resumo_pct": oee_proxy,
        },
        "semanal": {
            "total_produzido": prod_total,
            "comparacao_maquinas": comparacao_maquinas,
            "comparacao_turnos": por_turno,
            "produtos_top": produtos_rank[:15],
            "tempo_parado_s": parada_s,
            "tempo_parado_fmt": _fmt_hms(parada_s),
            "eficiencia_proxy_pct": oee_proxy,
        },
        "mensal": {
            "producao_total": prod_total,
            "ranking_maquinas": comparacao_maquinas,
            "ranking_operadores": por_operador[:20],
            "oee_medio_proxy_pct": oee_proxy,
            "perdas_estimadas": max(0, perdas_estimadas),
            "produtividade_un_h": round(prod_total / max(0.01, span_s / 3600.0), 1),
        },
        "paradas_lista": paradas[:500],
        "detalhado": detalhado[:2000],
    }


def _empty_payload(flt: FiltrosRelatorio, msg: str) -> dict[str, Any]:
    return {
        "ok": True,
        "aviso": msg,
        "filtros": {
            "inicio": flt.inicio.date().isoformat(),
            "fim": flt.fim.date().isoformat(),
            "maquina_id": flt.maquina_id,
            "operador": flt.operador,
            "turno": flt.turno,
            "produto": flt.produto,
            "setor": flt.setor,
        },
        "notas_metodologia": [],
        "kpis": {
            "producao_total": 0,
            "total_apontamentos_db": 0,
            "eventos_producao": 0,
            "tempo_parado_total_s": 0,
            "tempo_parado_fmt": "00:00:00",
            "disponibilidade_proxy_pct": 0.0,
            "oee_proxy_pct": 0.0,
            "perdas_estimadas_vs_meta": 0,
            "maquinas_no_escopo": 0,
        },
        "diario": {
            "por_maquina": [],
            "por_operador": [],
            "por_turno": [],
            "produtos": [],
            "paradas_motivo": [],
            "eficiencia_resumo_pct": 0.0,
        },
        "semanal": {
            "total_produzido": 0,
            "comparacao_maquinas": [],
            "comparacao_turnos": [],
            "produtos_top": [],
            "tempo_parado_s": 0,
            "tempo_parado_fmt": "00:00:00",
            "eficiencia_proxy_pct": 0.0,
        },
        "mensal": {
            "producao_total": 0,
            "ranking_maquinas": [],
            "ranking_operadores": [],
            "oee_medio_proxy_pct": 0.0,
            "perdas_estimadas": 0,
            "produtividade_un_h": 0.0,
        },
        "paradas_lista": [],
        "detalhado": [],
    }


def meta_filtros() -> dict[str, Any]:
    maquinas = []
    setores: set[str] = set()
    for mid, m in sorted(dados_maquinas.items()):
        if m.get("ativo", True) is False:
            continue
        st = str(m.get("setor") or "").strip()
        if st:
            setores.add(st)
        maquinas.append(
            {
                "id": int(mid),
                "nome": _nome_maquina(int(mid)),
                "setor": st,
            }
        )
    turnos: set[str] = set()
    operadores: set[str] = set()
    produtos_samples: set[str] = set()
    try:
        with SessionLocal() as session:
            for tu in session.scalars(select(Apontamento.turno).distinct()).all():
                s = str(tu or "").strip()
                if s:
                    turnos.add(s)
            for op in session.scalars(select(Apontamento.operador).distinct()).all():
                s = str(op or "").strip()
                if s:
                    operadores.add(s)
            for pr in session.scalars(select(Apontamento.produto).distinct().limit(200)).all():
                s = str(pr or "").strip()
                if s:
                    produtos_samples.add(s)
    except Exception:
        pass
    from services.config_params_db import list_operators

    for o in list_operators():
        n = str(o.get("nome") or "").strip()
        if n:
            operadores.add(n)
        tp = str(o.get("turno_padrao") or "").strip()
        if tp:
            turnos.add(tp)
    return {
        "ok": True,
        "maquinas": maquinas,
        "setores": sorted(setores),
        "turnos": sorted(turnos),
        "operadores": sorted(operadores),
        "produtos_amostra": sorted(produtos_samples)[:80],
    }


def build_workbook_bytes(payload: dict[str, Any], titulo_planilha: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Resumo"
    hdr_fill = PatternFill("solid", fgColor="1A4A62")
    hdr_font = Font(color="FFFFFF", bold=True)
    row = 1
    ws0.cell(row, 1, titulo_planilha)
    ws0.cell(row, 1).font = Font(bold=True, size=14)
    row += 1
    ws0.cell(row, 1, "Gerado em")
    ws0.cell(row, 2, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row += 2
    kp = payload.get("kpis") or {}
    for k, v in kp.items():
        ws0.cell(row, 1, k)
        ws0.cell(row, 2, v)
        row += 1

    def sheet_table(name: str, headers: list[str], data_rows: list[list[Any]]):
        ws = wb.create_sheet(title=name[:31])
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        for r, line in enumerate(data_rows, 2):
            for c, val in enumerate(line, 1):
                ws.cell(r, c, val)
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18
        return ws

    sem = payload.get("semanal") or {}
    sheet_table(
        "Por máquina",
        ["Máquina", "Setor", "Produção", "Tempo parado (s)", "Eficiência est. %"],
        [
            [
                x.get("nome"),
                x.get("setor"),
                x.get("producao"),
                x.get("tempo_parado_s"),
                x.get("eficiencia_est"),
            ]
            for x in (sem.get("comparacao_maquinas") or [])
        ],
    )
    sheet_table(
        "Por operador",
        ["Operador", "Quantidade"],
        [[x.get("nome"), x.get("quantidade")] for x in (payload.get("diario") or {}).get("por_operador") or []],
    )
    sheet_table(
        "Por turno",
        ["Turno", "Quantidade"],
        [[x.get("nome"), x.get("quantidade")] for x in (payload.get("diario") or {}).get("por_turno") or []],
    )
    det = payload.get("detalhado") or []
    wsd = sheet_table(
        "Detalhado",
        [
            "Produto",
            "Operador",
            "Turno",
            "Máquina",
            "Qtd",
            "Horário",
            "Tempo prod. (est.)",
            "Tempo parado",
            "Observações",
        ],
        [
            [
                d.get("produto"),
                d.get("operador"),
                d.get("turno"),
                d.get("maquina"),
                d.get("quantidade"),
                d.get("horario"),
                d.get("tempo_producao"),
                d.get("tempo_parado"),
                d.get("observacoes"),
            ]
            for d in det
        ],
    )
    lr = len(det) + 1
    if lr > 1:
        wsd.auto_filter.ref = f"A1:I{lr}"
    par = payload.get("paradas_lista") or []
    sheet_table(
        "Paradas",
        ["Máquina", "Início", "Retorno", "Duração (s)", "Motivo", "Operador", "Turno"],
        [
            [
                p.get("maquina_nome"),
                p.get("inicio"),
                p.get("retorno"),
                p.get("duracao_s"),
                p.get("motivo"),
                p.get("operador"),
                p.get("turno"),
            ]
            for p in par
        ],
    )
    ws0.freeze_panes = "A4"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _resolve_logo_path() -> Path | None:
    from services.runtime_config import visual_branding

    b = visual_branding()
    url = str(b.get("logo_url") or "").strip()
    if url.startswith("/static/"):
        p = _ROOT / url.lstrip("/").replace("/", os.sep)
        if p.is_file():
            return p
    p = _ROOT / "logo.png"
    if p.is_file():
        return p
    return None


def _pdf_fmt_br_int(n: Any) -> str:
    try:
        v = int(n)
        return f"{v:,}".replace(",", ".")
    except (TypeError, ValueError):
        return "—"


class _PdfDrawingFlowable(Flowable):
    """Flowable que desenha um reportlab.graphics.shapes.Drawing."""

    __slots__ = ("drawing", "width", "height")

    def __init__(self, drawing: Any) -> None:
        super().__init__()
        self.drawing = drawing
        self.width = float(getattr(drawing, "width", 400))
        self.height = float(getattr(drawing, "height", 160))

    def draw(self) -> None:
        from reportlab.graphics import renderPDF

        renderPDF.draw(self.drawing, self.canv, 0, 0)

    def wrap(self, availWidth: float, availHeight: float) -> tuple[float, float]:
        return self.width, self.height

    def split(self, availWidth: float, availHeight: float) -> list:
        return []


def _pdf_vertical_bars(labels: list[str], values: list[float], title: str, w_pt: float = 480, h_pt: float = 175) -> Any:
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors as rl_colors

    labels = [str(x)[:18] for x in labels]
    vals = [float(v) if v is not None else 0.0 for v in values]
    if not vals:
        vals = [0.0]
    d = Drawing(w_pt, h_pt + 22)
    d.add(String(8, h_pt + 6, title, fontSize=9, fillColor=rl_colors.HexColor("#0f2744"), fontName="Helvetica-Bold"))
    bc = VerticalBarChart()
    bc.x = 40
    bc.y = 28
    bc.height = h_pt - 38
    bc.width = w_pt - 52
    bc.data = [vals]
    bc.strokeColor = rl_colors.HexColor("#cbd5e1")
    bc.bars[0].fillColor = rl_colors.HexColor("#2d7ab8")
    n = max(len(vals), 1)
    bw = (bc.width / max(n * 1.35, 1)) if n else 10
    bc.barWidth = min(18, max(6, bw))
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.strokeColor = rl_colors.HexColor("#94a3b8")
    bc.categoryAxis.strokeWidth = 0.5
    bc.categoryAxis.labels.fontName = "Helvetica"
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.angle = 30
    bc.valueAxis.strokeColor = rl_colors.HexColor("#cbd5e1")
    bc.valueAxis.labels.fontName = "Helvetica"
    bc.valueAxis.labels.fontSize = 7
    bc.valueAxis.valueMin = 0
    bc.groupSpacing = 6
    d.add(bc)
    return d


def _pdf_horizontal_bars(labels: list[str], values: list[float], title: str, w_pt: float = 480, row_h: float = 16) -> Any:
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.lib import colors as rl_colors

    navy = rl_colors.HexColor("#1e4976")
    muted = rl_colors.HexColor("#475569")
    labels = [str(x)[:32] for x in labels]
    vals = [float(v) if v is not None else 0.0 for v in values]
    if not vals:
        vals = [0.0]
        labels = ["—"]
    mx = max(vals) or 1.0
    head = 22
    h_pt = head + len(labels) * row_h + 14
    d = Drawing(w_pt, h_pt)
    d.add(String(8, h_pt - 12, title, fontSize=9, fillColor=rl_colors.HexColor("#0f2744"), fontName="Helvetica-Bold"))
    bar_max = w_pt - 150
    x0 = 130
    y_top = h_pt - head - 8
    for i, (lb, v) in enumerate(zip(labels, vals)):
        y = y_top - i * row_h
        bw = bar_max * (v / mx)
        d.add(String(4, y - 2, lb, fontSize=7.5, fillColor=muted, fontName="Helvetica"))
        d.add(Rect(x0, y - 10, bw, 9, fillColor=navy, strokeColor=rl_colors.HexColor("#0c2744"), strokeWidth=0.3))
        d.add(String(x0 + bar_max + 4, y - 2, _pdf_fmt_br_int(int(v)) if v == int(v) else f"{v:.1f}", fontSize=7, fillColor=muted))
    return d


def _pdf_make_header_table(
    logo_p: Path | None,
    titulo_pdf: str,
    periodo_txt: str,
    gerado_txt: str,
    empresa: str,
    status_txt: str,
    styles: dict[str, Any],
    tw: float,
) -> Any:
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Image, Paragraph, Table, TableStyle

    navy = rl_colors.HexColor("#0c2744")

    left_cells: list[Any] = []
    if logo_p and logo_p.suffix.lower() in (".png", ".jpg", ".jpeg"):
        try:
            left_cells.append(Image(str(logo_p), width=3.4 * 28.346, height=1.25 * 28.346))
        except Exception:
            left_cells.append(Paragraph("", styles["hdr_white"]))
    else:
        left_cells.append(Paragraph("", styles["hdr_white"]))
    left_cells.append(
        Paragraph(
            f"<font size=18><b>{empresa}</b></font><br/><font size=11>MES · Relatório executivo industrial</font>",
            styles["hdr_white"],
        )
    )

    meta_html = (
        f"<font size=9><b>{titulo_pdf}</b></font><br/><br/>"
        f"<font size=8>Período analisado<br/><b>{periodo_txt}</b></font><br/><br/>"
        f"<font size=8>Emissão<br/><b>{gerado_txt}</b></font><br/><br/>"
        f"<font size=8>Status<br/><b>{status_txt}</b></font>"
    )
    right_cell = Paragraph(meta_html, styles["hdr_white_right"])

    inner = Table([[left_cells[0], left_cells[1]]], colWidths=[4 * 28.346 + 10, tw - (4 * 28.346 + 10) - 160])
    inner.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    meta_tbl = Table([[inner, right_cell]], colWidths=[tw - 175, 165])
    meta_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), navy),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 16),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
                ("LINEBELOW", (0, 0), (-1, -1), 1.5, rl_colors.HexColor("#38bdf8")),
            ]
        )
    )
    return meta_tbl


def _pdf_kpi_cards(payload: dict[str, Any], styles: dict[str, Any], tw: float) -> Any:
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    kp = payload.get("kpis") or {}
    sem = payload.get("semanal") or {}
    dia = payload.get("diario") or {}
    cmq = sem.get("comparacao_maquinas") or []
    po = dia.get("por_operador") or []

    best_m_name = str((cmq[0] or {}).get("nome") or "—") if cmq else "—"
    best_m_val = _pdf_fmt_br_int((cmq[0] or {}).get("producao")) if cmq else "—"

    best_o_name = str((po[0] or {}).get("nome") or "—") if po else "—"
    best_o_val = _pdf_fmt_br_int((po[0] or {}).get("quantidade")) if po else "—"

    prod = _pdf_fmt_br_int(kp.get("producao_total"))
    oee = kp.get("oee_proxy_pct")
    oee_s = f"{oee} %" if oee is not None else "—"
    disp = kp.get("disponibilidade_proxy_pct")
    disp_s = f"{disp} %" if disp is not None else "—"
    par = str(kp.get("tempo_parado_fmt") or "—")

    def card(label: str, value: str, sub: str = "") -> Any:
        lab = Paragraph(f"<font size=7 color='#64748b'>{label}</font>", styles["card_lab"])
        val = Paragraph(f"<font size=16><b>{value}</b></font>", styles["card_val"])
        rows: list[list[Any]] = [[lab], [val]]
        if sub:
            rows.append([Paragraph(f"<font size=9 color='#475569'>{sub}</font>", styles["card_lab"])])
        inner = Table(rows, colWidths=[tw / 3 - 18])
        inner.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), rl_colors.HexColor("#f4f7fb")),
                    ("BOX", (0, 0), (-1, -1), 0.9, rl_colors.HexColor("#cbd5e1")),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        return inner

    w3 = tw / 3 - 12
    r1 = Table(
        [
            [
                card("Produção total", prod, ""),
                card("OEE (proxy)", oee_s, ""),
                card("Disponibilidade", disp_s, ""),
            ]
        ],
        colWidths=[w3, w3, w3],
    )
    r1.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))

    r2 = Table(
        [
            [
                card("Tempo parado (total)", par, ""),
                card("Melhor máquina", best_m_name[:20], best_m_val + " un." if best_m_val != "—" else "—"),
                card("Melhor operador", best_o_name[:20], best_o_val + " un." if best_o_val != "—" else "—"),
            ]
        ],
        colWidths=[w3, w3, w3],
    )
    r2.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("TOPPADDING", (0, 0), (-1, -1), 8), ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))

    outer = Table([[r1], [r2]], colWidths=[tw])
    outer.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("TOPPADDING", (0, 0), (-1, -1), 0)]))
    return outer


def _pdf_ent_table(headers: list[str], rows: list[list[str]], col_widths: list[float]) -> Any:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Table, TableStyle

    navy = rl_colors.HexColor("#0c2744")
    hdr_style = ParagraphStyle(
        name="pdf_ent_th",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=11,
        textColor=rl_colors.HexColor("#f8fafc"),
        alignment=1,
    )
    cell_style = ParagraphStyle(
        name="pdf_ent_td",
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=rl_colors.HexColor("#334155"),
    )
    hdr_par = [Paragraph(str(h), hdr_style) for h in headers]
    body = []
    for r in rows:
        body.append([Paragraph(str(cell), cell_style) for cell in r])
    data = [hdr_par] + body
    t = Table(data, colWidths=col_widths)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), navy),
                ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.HexColor("#f8fafc")),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.HexColor("#ffffff"), rl_colors.HexColor("#f1f5f9")]),
                ("GRID", (0, 0), (-1, -1), 0.35, rl_colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return t


def build_pdf_bytes(payload: dict[str, Any], titulo: str) -> bytes:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.4 * cm,
        leftMargin=1.4 * cm,
        topMargin=1.1 * cm,
        bottomMargin=1.35 * cm,
    )
    tw = A4[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="HdrWhite", parent=styles["Normal"], textColor=rl_colors.HexColor("#f8fafc"), fontName="Helvetica", leading=13))
    styles.add(
        ParagraphStyle(
            name="HdrWhiteRight",
            parent=styles["Normal"],
            textColor=rl_colors.HexColor("#e2e8f0"),
            fontName="Helvetica",
            alignment=2,
            leading=12,
        )
    )
    styles.add(ParagraphStyle(name="SecTitle", parent=styles["Heading2"], textColor=rl_colors.HexColor("#0c2744"), fontSize=12, spaceAfter=8, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="card_lab", parent=styles["Normal"], alignment=0, leading=10))
    styles.add(ParagraphStyle(name="card_val", parent=styles["Normal"], alignment=0, leading=20))

    hdr_map = {
        "hdr_white": styles["HdrWhite"],
        "hdr_white_right": styles["HdrWhiteRight"],
        "sec": styles["SecTitle"],
        "card_lab": styles["card_lab"],
        "card_val": styles["card_val"],
    }

    logo_p = _resolve_logo_path()
    from services.runtime_config import visual_branding

    empresa = str(visual_branding().get("nome_empresa") or "INDUPACK")
    fl = payload.get("filtros") or {}
    ini = fl.get("inicio") or "—"
    fim = fl.get("fim") or "—"
    periodo_txt = f"{ini} — {fim}"
    now = datetime.now()
    gerado_txt = now.strftime("%d/%m/%Y   %H:%M")
    ok = payload.get("ok", True)
    aviso = str(payload.get("aviso") or "").strip()
    status_txt = "Consolidado" if ok and not aviso else ("Atenção — " + aviso[:80] if aviso else "Consolidado")

    story: list[Any] = []
    story.append(
        _pdf_make_header_table(
            logo_p,
            titulo,
            periodo_txt,
            gerado_txt,
            empresa,
            status_txt,
            hdr_map,
            tw,
        )
    )
    story.append(Spacer(1, 0.55 * cm))
    story.append(Paragraph("<b>Indicadores executivos</b>", hdr_map["sec"]))
    story.append(Spacer(1, 0.25 * cm))
    story.append(_pdf_kpi_cards(payload, hdr_map, tw))
    story.append(Spacer(1, 0.55 * cm))

    sem = payload.get("semanal") or {}
    cmq = sem.get("comparacao_maquinas") or []
    if cmq:
        labs = [str(x.get("nome") or "")[:16] for x in cmq[:10]]
        vals = [float(x.get("producao") or 0) for x in cmq[:10]]
        story.append(Paragraph("<b>Produção por máquina</b>", hdr_map["sec"]))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_PdfDrawingFlowable(_pdf_vertical_bars(labs, vals, "Volume produzido no período (unidades)", tw, 168)))
        story.append(Spacer(1, 0.35 * cm))
        rows_m = []
        for x in cmq[:18]:
            rows_m.append(
                [
                    str(x.get("nome") or "—")[:26],
                    str(x.get("setor") or "—")[:18],
                    _pdf_fmt_br_int(x.get("producao")),
                    str(x.get("tempo_parado_fmt") or "—"),
                    f"{x.get('eficiencia_est')}%",
                ]
            )
        story.append(
            _pdf_ent_table(
                ["Máquina", "Setor", "Produção", "Tempo parado", "Eficiência est."],
                rows_m,
                [tw * 0.26, tw * 0.18, tw * 0.18, tw * 0.20, tw * 0.18],
            )
        )
        story.append(Spacer(1, 0.45 * cm))

    prod_top = sem.get("produtos_top") or []
    if prod_top:
        labs_p = [str(x.get("nome") or "")[:18] for x in prod_top[:8]]
        vals_p = [float(x.get("quantidade") or 0) for x in prod_top[:8]]
        story.append(Paragraph("<b>Distribuição — produtos em destaque</b>", hdr_map["sec"]))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_PdfDrawingFlowable(_pdf_vertical_bars(labs_p, vals_p, "Quantidade por produto (unidades)", tw, 168)))
        story.append(Spacer(1, 0.35 * cm))

    if cmq:
        labs_e = [str(x.get("nome") or "")[:14] for x in cmq[:10]]
        vals_e = [float(x.get("eficiencia_est") or 0) for x in cmq[:10]]
        story.append(Paragraph("<b>Eficiência estimada por máquina</b>", hdr_map["sec"]))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_PdfDrawingFlowable(_pdf_vertical_bars(labs_e, vals_e, "Eficiência vs meta referência (%)", tw, 155)))

    dia = payload.get("diario") or {}
    par_mot = dia.get("paradas_motivo") or []
    if par_mot:
        story.append(Spacer(1, 0.45 * cm))
        story.append(Paragraph("<b>Paradas — tempo por motivo</b>", hdr_map["sec"]))
        story.append(Spacer(1, 0.15 * cm))
        labs_s = [str(x.get("motivo") or x.get("nome") or "") for x in par_mot[:12]]
        vals_s = [float(x.get("duracao_s") or 0) for x in par_mot[:12]]
        story.append(_PdfDrawingFlowable(_pdf_horizontal_bars(labs_s, vals_s, "Somatório de duração (segundos)", tw)))
        story.append(Spacer(1, 0.25 * cm))
        rows_p = [[str(x.get("motivo") or "—")[:36], str(x.get("duracao_fmt") or "—")] for x in par_mot[:14]]
        story.append(_pdf_ent_table(["Motivo / tipo", "Tempo parado"], rows_p, [tw * 0.72, tw * 0.28]))

    det = (payload.get("detalhado") or [])[:35]
    if det:
        story.append(Spacer(1, 0.45 * cm))
        story.append(Paragraph("<b>Apontamentos — amostra analítica</b>", hdr_map["sec"]))
        story.append(Spacer(1, 0.15 * cm))
        rows_d = []
        for d in det:
            rows_d.append(
                [
                    str(d.get("produto") or "")[:22],
                    str(d.get("operador") or "")[:14],
                    str(d.get("maquina") or "")[:14],
                    _pdf_fmt_br_int(d.get("quantidade")),
                    str(d.get("horario") or "")[:16],
                ]
            )
        story.append(_pdf_ent_table(["Produto", "Operador", "Máquina", "Quantidade", "Horário"], rows_d, [tw * 0.30, tw * 0.18, tw * 0.18, tw * 0.14, tw * 0.20]))

    notas = payload.get("notas_metodologia") or []
    if notas:
        story.append(Spacer(1, 0.45 * cm))
        story.append(Paragraph("<b>Metodologia e observações</b>", hdr_map["sec"]))
        for line in notas[:6]:
            story.append(Paragraph(f"<font size=8 color='#64748b'>• {line}</font>", styles["Normal"]))

    emit_iso = now.strftime("%Y-%m-%d %H:%M:%S")

    def _footer(canvas: Any, doc: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(rl_colors.HexColor("#cbd5e1"))
        canvas.setLineWidth(0.4)
        yl = 0.95 * cm
        canvas.line(doc.leftMargin, yl + 10, doc.pagesize[0] - doc.rightMargin, yl + 10)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(rl_colors.HexColor("#64748b"))
        left = f"INDUPACK MES · {empresa}"
        right = f"Emitido em {emit_iso} · pág. {canvas.getPageNumber()}"
        canvas.drawString(doc.leftMargin, yl, left)
        canvas.drawRightString(doc.pagesize[0] - doc.rightMargin, yl, right)
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    return buf.getvalue()


def enviar_pdf_email(destinatario: str, assunto: str, corpo: str, pdf_bytes: bytes, filename: str) -> dict[str, Any]:
    host = (os.environ.get("INDUPACK_SMTP_HOST") or "").strip()
    port = int(os.environ.get("INDUPACK_SMTP_PORT") or "587")
    user = (os.environ.get("INDUPACK_SMTP_USER") or "").strip()
    pw = (os.environ.get("INDUPACK_SMTP_PASSWORD") or "").strip()
    from_addr = (os.environ.get("INDUPACK_SMTP_FROM") or user or "").strip()
    if not host or not from_addr:
        return {
            "ok": False,
            "erro": "smtp_nao_configurado",
            "mensagem": "O envio por e-mail ainda não está disponível. Solicite ao administrador do sistema a configuração do serviço de notificações.",
        }
    import smtplib
    from email.message import EmailMessage

    msg = EmailMessage()
    msg["Subject"] = assunto or "Relatório INDUPACK"
    msg["From"] = from_addr
    msg["To"] = destinatario
    msg.set_content(corpo or "Segue em anexo o relatório gerado pelo INDUPACK.")
    msg.add_attachment(
        pdf_bytes,
        main="application",
        subtype="pdf",
        filename=filename or "relatorio_indupack.pdf",
    )
    try:
        with smtplib.SMTP(host, port, timeout=25) as smtp:
            smtp.ehlo()
            try:
                smtp.starttls()
            except Exception:
                pass
            if user:
                smtp.login(user, pw)
            smtp.send_message(msg)
        return {"ok": True, "mensagem": "Relatório enviado por e-mail com sucesso."}
    except Exception as exc:
        _log.warning("Falha ao enviar relatório por e-mail", exc_info=True)
        return {
            "ok": False,
            "erro": "smtp_falha",
            "mensagem": "Não foi possível enviar o relatório no momento. Tente novamente em instantes ou contate o administrador do sistema.",
        }


def stub_agendamento_relatorios() -> dict[str, Any]:
    """Estrutura futura: envio automático diário / semanal / mensal (persistir em system_settings)."""
    return {
        "automatico_diario": {"habilitado": False, "horario": "07:30", "destinatarios": []},
        "automatico_semanal": {"habilitado": False, "dia_semana": 1, "horario": "08:00", "destinatarios": []},
        "automatico_mensal": {"habilitado": False, "dia_mes": 1, "horario": "08:30", "destinatarios": []},
        "nota": "Persistência e scheduler podem usar os mesmos endpoints de exportação + enviar_pdf_email.",
    }
